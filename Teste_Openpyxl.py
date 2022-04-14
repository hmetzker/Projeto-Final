import pandas as pd
from openpyxl import Workbook, load_workbook

p_csv = pd.read_csv('BD.csv', sep=';')
print(p_csv)

p_xlsx = pd.read_excel('BD.xlsx')
print()
print(p_xlsx)

p_xlsx = load_workbook('BD.xlsx')
p_ativa = p_xlsx.active
max_linha = p_ativa.max_row
max_coluna = p_ativa.max_column

print()
for i in range(1, max_linha+1):
    for j in range(1, max_coluna+1):
        celula = p_ativa.cell(row=i, column=j)
        if i==1:
            print(celula.value, end=' ')
        else:
            print(celula.value, end=' ')
    print('\n')
